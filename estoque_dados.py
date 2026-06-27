"""
estoque_dados.py

Camada de dados do Sistema de Controle de Estoque. Todas as funções que
leem ou escrevem no arquivo estoque.xlsx ficam centralizadas aqui — o
resto do sistema (menu, relatórios, etc.) nunca deve abrir o Excel
diretamente.

Convenções:
    - Cada produto é representado como um dicionário Python com as
      mesmas chaves das colunas da aba Produtos.
    - Cada movimento é representado como um dicionário com as chaves
      das colunas da aba Movimentos.
    - Datas de movimento são armazenadas como objetos datetime.date.
"""

from datetime import date, datetime
from openpyxl import load_workbook

NOME_ARQUIVO = "estoque.xlsx"

COLUNAS_PRODUTOS = [
    "ID",
    "Nome",
    "Categoria",
    "Local",
    "Quantidade Atual",
    "Ponto de Reposição",
]

COLUNAS_MOVIMENTOS = [
    "Data",
    "ID Produto",
    "Tipo",
    "Quantidade",
]


class ArquivoNaoEncontradoError(Exception):
    """Levantado quando estoque.xlsx não existe (rode setup_planilha.py)."""


class ProdutoNaoEncontradoError(Exception):
    """Levantado quando um ID de produto não existe na aba Produtos."""


def _abrir_workbook(caminho: str = NOME_ARQUIVO):
    """Abre o workbook, com mensagem de erro clara se o arquivo não existir."""
    try:
        return load_workbook(caminho)
    except FileNotFoundError:
        raise ArquivoNaoEncontradoError(
            f"O arquivo '{caminho}' não foi encontrado. "
            "Rode 'python setup_planilha.py' primeiro."
        )
    except PermissionError:
        raise PermissionError(
            f"Não foi possível abrir '{caminho}'. "
            "Verifique se ele não está aberto no Excel e tente novamente."
        )


def _linha_para_dict(colunas, linha):
    """Converte uma linha (tupla de células) num dicionário usando 'colunas'."""
    return {coluna: celula for coluna, celula in zip(colunas, linha)}


def carregar_produtos(caminho: str = NOME_ARQUIVO) -> list[dict]:
    """Retorna todos os produtos da aba Produtos como lista de dicionários."""
    wb = _abrir_workbook(caminho)
    aba = wb["Produtos"]
    produtos = []
    for linha in aba.iter_rows(min_row=2, values_only=True):
        if linha[0] is None:  # pula linhas vazias no final
            continue
        produtos.append(_linha_para_dict(COLUNAS_PRODUTOS, linha))
    return produtos


def carregar_movimentos(caminho: str = NOME_ARQUIVO) -> list[dict]:
    """Retorna todos os movimentos da aba Movimentos como lista de dicionários."""
    wb = _abrir_workbook(caminho)
    aba = wb["Movimentos"]
    movimentos = []
    for linha in aba.iter_rows(min_row=2, values_only=True):
        if linha[0] is None:
            continue
        mov = _linha_para_dict(COLUNAS_MOVIMENTOS, linha)
        # Normaliza a data: openpyxl pode retornar datetime ou date
        if isinstance(mov["Data"], datetime):
            mov["Data"] = mov["Data"].date()
        movimentos.append(mov)
    return movimentos


def buscar_produto(id_produto: str, caminho: str = NOME_ARQUIVO) -> dict:
    """Retorna o produto com o ID informado.

    Levanta ProdutoNaoEncontradoError se não existir.
    """
    for produto in carregar_produtos(caminho):
        if produto["ID"] == id_produto:
            return produto
    raise ProdutoNaoEncontradoError(f"Produto '{id_produto}' não encontrado.")


def produto_existe(id_produto: str, caminho: str = NOME_ARQUIVO) -> bool:
    """Retorna True se o ID de produto já existe na aba Produtos."""
    try:
        buscar_produto(id_produto, caminho)
        return True
    except ProdutoNaoEncontradoError:
        return False


def salvar_produto(produto: dict, caminho: str = NOME_ARQUIVO) -> None:
    """Insere um novo produto ou atualiza um existente (mesmo ID).

    'produto' deve ser um dicionário com todas as chaves de
    COLUNAS_PRODUTOS.
    """
    wb = _abrir_workbook(caminho)
    aba = wb["Produtos"]

    linha_existente = None
    for idx, linha in enumerate(
        aba.iter_rows(min_row=2, max_col=1, values_only=True), start=2
    ):
        if linha[0] == produto["ID"]:
            linha_existente = idx
            break

    valores = [produto[coluna] for coluna in COLUNAS_PRODUTOS]

    if linha_existente:
        for col_idx, valor in enumerate(valores, start=1):
            aba.cell(row=linha_existente, column=col_idx, value=valor)
    else:
        aba.append(valores)

    wb.save(caminho)


def adicionar_movimento(
    id_produto: str,
    tipo: str,
    quantidade: float,
    data_movimento: date = None,
    caminho: str = NOME_ARQUIVO,
) -> None:
    """Grava uma nova linha na aba Movimentos.

    'tipo' deve ser "Entrada" ou "Saída". Não recalcula a quantidade do
    produto — chame recalcular_quantidade() separadamente após gravar.
    """
    if tipo not in ("Entrada", "Saída"):
        raise ValueError("tipo deve ser 'Entrada' ou 'Saída'")

    if data_movimento is None:
        data_movimento = date.today()

    wb = _abrir_workbook(caminho)
    aba = wb["Movimentos"]
    aba.append([data_movimento, id_produto, tipo, quantidade])
    wb.save(caminho)


def recalcular_quantidade(id_produto: str, caminho: str = NOME_ARQUIVO) -> float:
    """Recalcula a Quantidade Atual de um produto a partir do histórico
    de Movimentos, e salva o resultado na aba Produtos.

    Retorna a nova quantidade calculada.
    """
    movimentos = carregar_movimentos(caminho)
    saldo = 0.0
    for mov in movimentos:
        if mov["ID Produto"] != id_produto:
            continue
        if mov["Tipo"] == "Entrada":
            saldo += mov["Quantidade"]
        else:  # Saída
            saldo -= mov["Quantidade"]

    produto = buscar_produto(id_produto, caminho)
    produto["Quantidade Atual"] = saldo
    salvar_produto(produto, caminho)
    return saldo
